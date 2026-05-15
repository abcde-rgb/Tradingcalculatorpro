"""missing_apis.py

All the missing/incomplete APIs from TradingCalculator PRO:

1. Forex real prices (ExchangeRate-API + yfinance fallback)
2. Indices real prices (yfinance)
3. Commodities real prices (yfinance GC=F, CL=F, SI=F, etc.)
4. Crypto OHLC for any symbol via yfinance (not only CoinGecko 11)
5. /auth/forgot-password  — sends reset token via email
6. /auth/reset-password   — validates token and sets new password
7. /auth/verify-email     — send + confirm email verification token
8. /subscriptions/change-plan — real Stripe proration upgrade/downgrade
9. Stripe webhook events: customer.subscription.deleted +
                          invoice.payment_failed
10. /api/performance/export  — CSV / Excel export of trades
11. /api/calculations/{id}/save-to-journal  — prefill journal from calculator
12. MongoDB unique sparse index on users.email (startup)
"""

from __future__ import annotations

import csv
import io
import logging
import os
import uuid
from datetime import datetime, timedelta, timezone
from typing import Any, Dict, List, Optional

import httpx
import stripe
from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import Response, StreamingResponse
from fastapi.security import HTTPAuthorizationCredentials, HTTPBearer
from pydantic import BaseModel, EmailStr

# ── These helpers are imported from the main server module at registration time
# (injected via register(router, db, helpers)).  We define placeholders here
# that get replaced by register().
db = None  # type: ignore[assignment]
require_user = None  # type: ignore[assignment]
check_premium = None  # type: ignore[assignment]
SUBSCRIPTION_PLANS: Dict[str, Any] = {}
hash_password = None  # type: ignore[assignment]
SENDGRID_API_KEY = ""
SENDER_EMAIL = ""
STRIPE_API_KEY = ""
get_setting = None  # type: ignore[assignment]
JWT_SECRET = ""
JWT_ALGORITHM = "HS256"

router = APIRouter()
security = HTTPBearer(auto_error=False)


# ─────────────────────────────────────────────────────────────────────
# Proxy dependencies — at decoration time `require_user` is None, so we
# use a proxy callable that defers resolution until request time.
# ─────────────────────────────────────────────────────────────────────
async def _require_user_proxy(
    credentials: HTTPAuthorizationCredentials = Depends(security),
) -> Dict[str, Any]:
    if require_user is None:
        raise HTTPException(status_code=503, detail="Service not initialized")
    return await require_user(credentials)

# ---------------------------------------------------------------------------
# 0.  STARTUP — MongoDB unique index on users.email
# ---------------------------------------------------------------------------

async def ensure_email_unique_index(database) -> None:
    """Create a unique sparse index on users.email to prevent race-condition
    duplicate registrations at the DB level.
    This is idempotent: safe to call on every startup.
    """
    try:
        await database.users.create_index(
            "email",
            unique=True,
            sparse=True,
            name="users_email_unique",
        )
        logging.info("✅ unique index on users.email ensured")
    except Exception as e:
        logging.error(f"Could not create unique index on users.email: {e}")


# ---------------------------------------------------------------------------
# 1.  FOREX REAL PRICES
# ---------------------------------------------------------------------------

FOREX_PAIRS = [
    "EURUSD", "GBPUSD", "USDJPY", "USDCHF",
    "AUDUSD", "USDCAD", "NZDUSD", "EURGBP",
    "EURJPY", "GBPJPY",
]

_FOREX_STATIC_FALLBACK: Dict[str, Dict[str, float]] = {
    "EURUSD": {"price": 1.0856, "change": 0.12, "source": "fallback"},
    "GBPUSD": {"price": 1.2734, "change": -0.08, "source": "fallback"},
    "USDJPY": {"price": 149.85, "change": 0.25, "source": "fallback"},
    "USDCHF": {"price": 0.8812, "change": -0.05, "source": "fallback"},
    "AUDUSD": {"price": 0.6542, "change": 0.18, "source": "fallback"},
    "USDCAD": {"price": 1.3564, "change": 0.03, "source": "fallback"},
    "NZDUSD": {"price": 0.5987, "change": 0.22, "source": "fallback"},
    "EURGBP": {"price": 0.8525, "change": 0.15, "source": "fallback"},
    "EURJPY": {"price": 162.68, "change": 0.38, "source": "fallback"},
    "GBPJPY": {"price": 190.78, "change": 0.17, "source": "fallback"},
}


async def _fetch_forex_exchangerate() -> Optional[Dict[str, Any]]:
    """Try ExchangeRate-API (free tier, no key for base USD)."""
    try:
        async with httpx.AsyncClient(timeout=8) as client:
            r = await client.get("https://open.er-api.com/v6/latest/USD")
            if r.status_code != 200:
                return None
            data = r.json()
            rates = data.get("rates", {})
            result: Dict[str, Any] = {}
            # We need USD-centric pairs and cross rates
            usd_eur = rates.get("EUR", 0)
            usd_gbp = rates.get("GBP", 0)
            usd_jpy = rates.get("JPY", 0)
            usd_chf = rates.get("CHF", 0)
            usd_aud = rates.get("AUD", 0)
            usd_cad = rates.get("CAD", 0)
            usd_nzd = rates.get("NZD", 0)

            def safe_div(a, b):
                return round(a / b, 5) if b else 0

            result["EURUSD"] = {"price": safe_div(1, usd_eur), "change": 0.0, "source": "exchangerate-api"}
            result["GBPUSD"] = {"price": safe_div(1, usd_gbp), "change": 0.0, "source": "exchangerate-api"}
            result["USDJPY"] = {"price": round(usd_jpy, 3), "change": 0.0, "source": "exchangerate-api"}
            result["USDCHF"] = {"price": round(usd_chf, 5), "change": 0.0, "source": "exchangerate-api"}
            result["AUDUSD"] = {"price": safe_div(1, usd_aud), "change": 0.0, "source": "exchangerate-api"}
            result["USDCAD"] = {"price": round(usd_cad, 5), "change": 0.0, "source": "exchangerate-api"}
            result["NZDUSD"] = {"price": safe_div(1, usd_nzd), "change": 0.0, "source": "exchangerate-api"}
            result["EURGBP"] = {"price": safe_div(usd_gbp, usd_eur) if usd_eur else 0, "change": 0.0, "source": "exchangerate-api"}
            eur_usd = safe_div(1, usd_eur)
            result["EURJPY"] = {"price": round(eur_usd * usd_jpy, 3), "change": 0.0, "source": "exchangerate-api"}
            gbp_usd = safe_div(1, usd_gbp)
            result["GBPJPY"] = {"price": round(gbp_usd * usd_jpy, 3), "change": 0.0, "source": "exchangerate-api"}
            return result
    except Exception as e:
        logging.warning(f"ExchangeRate-API forex fetch failed: {e}")
        return None


async def _fetch_forex_yfinance() -> Optional[Dict[str, Any]]:
    """Fallback: use yfinance =X suffix pairs."""
    try:
        import yfinance as yf
        yf_pairs = {
            "EURUSD": "EURUSD=X", "GBPUSD": "GBPUSD=X", "USDJPY": "USDJPY=X",
            "USDCHF": "USDCHF=X", "AUDUSD": "AUDUSD=X", "USDCAD": "USDCAD=X",
            "NZDUSD": "NZDUSD=X", "EURGBP": "EURGBP=X", "EURJPY": "EURJPY=X",
            "GBPJPY": "GBPJPY=X",
        }
        symbols = " ".join(yf_pairs.values())
        tickers = yf.download(symbols, period="2d", interval="1h", progress=False, auto_adjust=True)
        result: Dict[str, Any] = {}
        close = tickers.get("Close", {})
        for pair, sym in yf_pairs.items():
            try:
                series = close[sym].dropna()
                if len(series) >= 2:
                    price = float(series.iloc[-1])
                    prev = float(series.iloc[-2])
                    change = round((price - prev) / prev * 100, 4) if prev else 0.0
                elif len(series) == 1:
                    price = float(series.iloc[-1])
                    change = 0.0
                else:
                    continue
                result[pair] = {"price": round(price, 5), "change": change, "source": "yfinance"}
            except Exception:
                continue
        return result if result else None
    except Exception as e:
        logging.warning(f"yfinance forex fetch failed: {e}")
        return None


@router.get("/forex-prices")
async def get_forex_prices_real():
    """Real forex prices: ExchangeRate-API → yfinance → static fallback."""
    result = await _fetch_forex_exchangerate()
    if result:
        return result
    result = await _fetch_forex_yfinance()
    if result:
        return result
    return _FOREX_STATIC_FALLBACK


# ---------------------------------------------------------------------------
# 2.  INDICES REAL PRICES
# ---------------------------------------------------------------------------

_INDICES_MAP: Dict[str, str] = {
    "SPX":  "^GSPC",
    "NDX":  "^NDX",
    "DJI":  "^DJI",
    "DAX":  "^GDAXI",
    "FTSE": "^FTSE",
    "N225": "^N225",
    "HSI":  "^HSI",
    "CAC":  "^FCHI",
    "VIX":  "^VIX",
    "RUT":  "^RUT",
}

_INDICES_STATIC_FALLBACK: Dict[str, Any] = {
    "SPX":  {"price": 5998.50, "change": 0.45, "source": "fallback"},
    "NDX":  {"price": 21245.80, "change": 0.68, "source": "fallback"},
    "DJI":  {"price": 44235.20, "change": 0.32, "source": "fallback"},
    "DAX":  {"price": 19785.60, "change": 0.28, "source": "fallback"},
    "FTSE": {"price": 8265.40, "change": 0.15, "source": "fallback"},
    "N225": {"price": 38542.80, "change": 0.52, "source": "fallback"},
    "HSI":  {"price": 19876.30, "change": -0.25, "source": "fallback"},
}


@router.get("/indices-prices")
async def get_indices_prices_real():
    """Real index prices from yfinance."""
    try:
        import yfinance as yf
        symbols = list(_INDICES_MAP.values())
        tickers = yf.download(" ".join(symbols), period="2d", interval="1d", progress=False, auto_adjust=True)
        close = tickers.get("Close", {})
        result: Dict[str, Any] = {}
        for label, sym in _INDICES_MAP.items():
            try:
                series = close[sym].dropna()
                if len(series) >= 2:
                    price = float(series.iloc[-1])
                    prev = float(series.iloc[-2])
                    change = round((price - prev) / prev * 100, 4) if prev else 0.0
                    result[label] = {"price": round(price, 2), "change": change, "source": "yfinance"}
                elif len(series) == 1:
                    result[label] = {"price": round(float(series.iloc[-1]), 2), "change": 0.0, "source": "yfinance"}
            except Exception:
                if label in _INDICES_STATIC_FALLBACK:
                    result[label] = _INDICES_STATIC_FALLBACK[label]
        if result:
            return result
    except Exception as e:
        logging.error(f"Indices yfinance error: {e}")
    return _INDICES_STATIC_FALLBACK


# ---------------------------------------------------------------------------
# 3.  COMMODITIES REAL PRICES
# ---------------------------------------------------------------------------

_COMMODITY_MAP: Dict[str, str] = {
    "gold":        "GC=F",
    "silver":      "SI=F",
    "crude_oil":   "CL=F",
    "brent":       "BZ=F",
    "natural_gas": "NG=F",
    "copper":      "HG=F",
    "platinum":    "PL=F",
    "wheat":       "ZW=F",
    "corn":        "ZC=F",
    "soybeans":    "ZS=F",
}

_COMMODITY_STATIC: Dict[str, Any] = {
    "gold":        {"usd": 2680, "eur": 2450, "usd_24h_change": 0.5, "source": "fallback"},
    "silver":      {"usd": 31.50, "eur": 28.80, "usd_24h_change": 0.8, "source": "fallback"},
    "crude_oil":   {"usd": 78.0, "eur": 71.5, "usd_24h_change": 0.3, "source": "fallback"},
    "brent":       {"usd": 82.0, "eur": 75.0, "usd_24h_change": 0.2, "source": "fallback"},
    "natural_gas": {"usd": 2.5, "eur": 2.3, "usd_24h_change": -0.5, "source": "fallback"},
    "copper":      {"usd": 4.2, "eur": 3.85, "usd_24h_change": 0.6, "source": "fallback"},
}


@router.get("/commodities-prices")
async def get_commodities_prices_real():
    """Real commodity prices from yfinance futures contracts."""
    try:
        import yfinance as yf
        # Fetch EUR/USD for conversion
        eur_usd = 0.917
        try:
            fx = yf.Ticker("EURUSD=X")
            hist = fx.history(period="2d")
            if not hist.empty:
                eur_usd = 1 / float(hist["Close"].iloc[-1])
        except Exception:
            pass

        result: Dict[str, Any] = {}
        for label, sym in _COMMODITY_MAP.items():
            try:
                t = yf.Ticker(sym)
                hist = t.history(period="2d")
                if hist.empty:
                    continue
                price = float(hist["Close"].iloc[-1])
                prev = float(hist["Close"].iloc[-2]) if len(hist) >= 2 else price
                change = round((price - prev) / prev * 100, 4) if prev else 0.0
                result[label] = {
                    "usd": round(price, 4),
                    "eur": round(price * eur_usd, 4),
                    "usd_24h_change": change,
                    "symbol": sym,
                    "source": "yfinance",
                }
            except Exception as e:
                logging.warning(f"Commodity {sym} fetch error: {e}")
                if label in _COMMODITY_STATIC:
                    result[label] = _COMMODITY_STATIC[label]
        return result if result else _COMMODITY_STATIC
    except Exception as e:
        logging.error(f"Commodities error: {e}")
        return _COMMODITY_STATIC


# ---------------------------------------------------------------------------
# 4.  CRYPTO OHLC — universal (any yfinance -USD symbol, not only 11 coins)
# ---------------------------------------------------------------------------

_COINGECKO_COIN_MAP: Dict[str, str] = {
    "BTC": "bitcoin", "ETH": "ethereum", "SOL": "solana", "BNB": "binancecoin",
    "XRP": "ripple", "ADA": "cardano", "DOGE": "dogecoin", "AVAX": "avalanche-2",
    "DOT": "polkadot", "LINK": "chainlink", "LTC": "litecoin",
}


def _ohlc_from_yfinance(symbol: str, days: int) -> Optional[List[Dict[str, Any]]]:
    """Fetch OHLC candles from yfinance for ANY asset symbol."""
    try:
        import yfinance as yf
        if days <= 7:
            interval = "1h"
        elif days <= 60:
            interval = "1d"
        else:
            interval = "1wk"
        period_str = f"{days}d" if days <= 730 else "2y"
        hist = yf.Ticker(symbol).history(period=period_str, interval=interval)
        if hist.empty:
            return None
        candles = []
        for idx, row in hist.iterrows():
            ts = int(idx.timestamp())
            candles.append({
                "time": ts,
                "open": round(float(row["Open"]), 6),
                "high": round(float(row["High"]), 6),
                "low":  round(float(row["Low"]), 6),
                "close": round(float(row["Close"]), 6),
                "volume": float(row.get("Volume", 0) or 0),
            })
        return candles
    except Exception as e:
        logging.warning(f"yfinance OHLC for {symbol}: {e}")
        return None


@router.get("/ohlc-universal/{symbol}")
async def get_ohlc_universal(symbol: str, days: int = 30) -> Dict[str, Any]:
    """
    Universal OHLC endpoint that works for ANY asset:
    - Crypto (BTC, ETH, SOL, MATIC, etc.) via CoinGecko first, yfinance fallback
    - Stocks / ETFs (AAPL, SPY, etc.) via yfinance
    - Forex (EURUSD=X) via yfinance
    - Commodities (GC=F) via yfinance
    - Indices (^GSPC) via yfinance
    """
    sym_upper = symbol.upper()
    empty: Dict[str, Any] = {"ohlc": [], "symbol": sym_upper, "source": "none"}

    # Step 1: Try CoinGecko for known crypto coins
    coin_id = _COINGECKO_COIN_MAP.get(sym_upper)
    if coin_id:
        try:
            async with httpx.AsyncClient(timeout=12) as client:
                r = await client.get(
                    f"https://api.coingecko.com/api/v3/coins/{coin_id}/market_chart",
                    params={"vs_currency": "usd", "days": days},
                )
            if r.status_code == 200:
                prices = r.json().get("prices", [])
                if prices:
                    interval_ms = (3600 * 1000) if days <= 7 else (4 * 3600 * 1000) if days <= 30 else (24 * 3600 * 1000)
                    ohlc: List[Dict[str, Any]] = []
                    bucket_start: Optional[int] = None
                    bucket_prices: List[float] = []
                    for ts, price in prices:
                        bucket = (int(ts) // interval_ms) * interval_ms
                        if bucket_start is None or bucket != bucket_start:
                            if bucket_prices and bucket_start is not None:
                                ohlc.append({"time": bucket_start // 1000, "open": bucket_prices[0],
                                             "high": max(bucket_prices), "low": min(bucket_prices),
                                             "close": bucket_prices[-1]})
                            bucket_start, bucket_prices = bucket, [price]
                        else:
                            bucket_prices.append(price)
                    if bucket_prices and bucket_start:
                        ohlc.append({"time": bucket_start // 1000, "open": bucket_prices[0],
                                     "high": max(bucket_prices), "low": min(bucket_prices),
                                     "close": bucket_prices[-1]})
                    return {"ohlc": ohlc, "symbol": sym_upper, "source": "coingecko"}
        except Exception:
            pass

    # Step 2: yfinance for everything (crypto as BTC-USD, stocks, forex, commodities)
    yf_sym = symbol  # Pass as-is (user may pass BTC-USD, AAPL, GC=F, etc.)
    # Also try appending -USD for bare crypto symbols not in CoinGecko map
    candles = _ohlc_from_yfinance(yf_sym, days)
    if candles:
        return {"ohlc": candles, "symbol": sym_upper, "source": "yfinance"}
    # Last attempt: try adding -USD suffix
    if not any(c in symbol for c in ["-", "=", "^", "."]):
        candles = _ohlc_from_yfinance(f"{sym_upper}-USD", days)
        if candles:
            return {"ohlc": candles, "symbol": sym_upper, "source": "yfinance_crypto"}

    return empty


# ---------------------------------------------------------------------------
# 5 + 6.  FORGOT PASSWORD / RESET PASSWORD
# ---------------------------------------------------------------------------

class ForgotPasswordRequest(BaseModel):
    email: EmailStr


class ResetPasswordRequest(BaseModel):
    token: str
    new_password: str


async def _send_reset_email(to_email: str, reset_url: str) -> bool:
    """Send password reset email via SendGrid."""
    if not SENDGRID_API_KEY:
        logging.warning("SendGrid not configured — reset email not sent")
        return False
    try:
        from sendgrid import SendGridAPIClient
        from sendgrid.helpers.mail import Mail
        msg = Mail(
            from_email=SENDER_EMAIL,
            to_emails=to_email,
            subject="🔐 Restablecer contraseña — Trading Calculator PRO",
            html_content=f"""
            <html><body style="font-family:Arial,sans-serif;background:#1a1a1a;color:#fff;padding:20px">
            <div style="max-width:600px;margin:0 auto;background:#2a2a2a;padding:30px;border-radius:10px">
                <h1 style="color:#00E676">Trading Calculator PRO</h1>
                <h2>Restablecer tu contraseña</h2>
                <p>Hemos recibido una solicitud para restablecer tu contraseña.</p>
                <p style="margin:30px 0">
                    <a href="{reset_url}" style="background:#00E676;color:#000;padding:14px 28px;
                       border-radius:8px;text-decoration:none;font-weight:bold;font-size:16px">
                        Restablecer contraseña
                    </a>
                </p>
                <p style="color:#888">Este enlace expira en 1 hora. Si no solicitaste esto, ignora este email.</p>
            </div>
            </body></html>
            """,
        )
        sg = SendGridAPIClient(SENDGRID_API_KEY)
        sg.send(msg)
        return True
    except Exception as e:
        logging.error(f"Reset email error: {e}")
        return False


@router.post("/auth/forgot-password")
async def forgot_password(request: Request, payload: ForgotPasswordRequest):
    """
    Generate a password-reset token, store it in MongoDB (expires in 1h),
    and send an email to the user with the reset link.
    Always returns 200 to avoid user-enumeration.
    """
    email_lc = payload.email.lower()
    user = await db.users.find_one({"email": email_lc}, {"_id": 0})
    if not user:
        # Avoid user enumeration: always return success
        return {"ok": True, "message": "Si existe una cuenta con ese email, recibirás un enlace de recuperación."}

    token = uuid.uuid4().hex + uuid.uuid4().hex  # 64-char random token
    expires_at = datetime.now(timezone.utc) + timedelta(hours=1)
    await db.password_reset_tokens.update_one(
        {"user_id": user["id"]},
        {"$set": {
            "user_id": user["id"],
            "email": email_lc,
            "token": token,
            "created_at": datetime.now(timezone.utc),
            "expires_at": expires_at,
            "used": False,
        }},
        upsert=True,
    )
    # Build reset URL from request origin
    origin = request.headers.get("origin") or request.headers.get("referer", "").rstrip("/")
    if not origin:
        origin = str(request.base_url).rstrip("/")
    reset_url = f"{origin}/reset-password?token={token}"

    sent = await _send_reset_email(email_lc, reset_url)
    resp: Dict[str, Any] = {
        "ok": True,
        "message": "Si existe una cuenta con ese email, recibirás un enlace de recuperación.",
    }
    # In development (no SendGrid), expose the token so devs can test
    if not sent:
        resp["dev_token"] = token
        resp["dev_reset_url"] = reset_url
    return resp


@router.post("/auth/reset-password")
async def reset_password(payload: ResetPasswordRequest):
    """
    Validate the token, set the new password, mark the token used,
    and revoke all existing sessions for that user.
    """
    if len(payload.new_password) < 6:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 6 caracteres")

    rec = await db.password_reset_tokens.find_one({"token": payload.token, "used": False})
    if not rec:
        raise HTTPException(status_code=400, detail="Token inválido o ya utilizado")

    expires_at = rec["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at.replace("Z", "+00:00"))
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=400, detail="El token ha expirado. Solicita uno nuevo.")

    import bcrypt
    new_hash = bcrypt.hashpw(payload.new_password.encode(), bcrypt.gensalt()).decode()
    await db.users.update_one(
        {"id": rec["user_id"]},
        {"$set": {
            "password": new_hash,
            "auth_provider": "password",
            "password_reset_at": datetime.now(timezone.utc).isoformat(),
        }},
    )
    await db.password_reset_tokens.update_one(
        {"token": payload.token},
        {"$set": {"used": True, "used_at": datetime.now(timezone.utc)}},
    )
    # Revoke all sessions for the user (user_revocations collection, same as admin reset)
    await db.user_revocations.update_one(
        {"user_id": rec["user_id"]},
        {"$set": {
            "user_id": rec["user_id"],
            "revoked_after": datetime.now(timezone.utc),
            "expires_at": datetime.now(timezone.utc) + timedelta(hours=25),
        }},
        upsert=True,
    )
    return {"ok": True, "message": "Contraseña actualizada. Por favor inicia sesión de nuevo."}


# ---------------------------------------------------------------------------
# 7.  EMAIL VERIFICATION
# ---------------------------------------------------------------------------

class VerifyEmailRequest(BaseModel):
    token: str


async def _send_verification_email(to_email: str, verify_url: str) -> bool:
    if not SENDGRID_API_KEY:
        return False
    try:
        from sendgrid import SendGridAPIClient
        from sendgrid.helpers.mail import Mail
        msg = Mail(
            from_email=SENDER_EMAIL,
            to_emails=to_email,
            subject="✅ Verifica tu email — Trading Calculator PRO",
            html_content=f"""
            <html><body style="font-family:Arial,sans-serif;background:#1a1a1a;color:#fff;padding:20px">
            <div style="max-width:600px;margin:0 auto;background:#2a2a2a;padding:30px;border-radius:10px">
                <h1 style="color:#00E676">Trading Calculator PRO</h1>
                <h2>Verifica tu dirección de email</h2>
                <p style="margin:30px 0">
                    <a href="{verify_url}" style="background:#00E676;color:#000;padding:14px 28px;
                       border-radius:8px;text-decoration:none;font-weight:bold">
                        Verificar email
                    </a>
                </p>
                <p style="color:#888">Este enlace expira en 24 horas.</p>
            </div>
            </body></html>
            """,
        )
        sg = SendGridAPIClient(SENDGRID_API_KEY)
        sg.send(msg)
        return True
    except Exception as e:
        logging.error(f"Verification email error: {e}")
        return False


@router.post("/auth/send-verification-email")
async def send_verification_email(request: Request, user: dict = Depends(_require_user_proxy)):
    """Send (or resend) an email-verification link for the logged-in user."""
    if user.get("email_verified"):
        return {"ok": True, "message": "Email ya verificado"}

    token = uuid.uuid4().hex + uuid.uuid4().hex
    expires_at = datetime.now(timezone.utc) + timedelta(hours=24)
    await db.email_verification_tokens.update_one(
        {"user_id": user["id"]},
        {"$set": {
            "user_id": user["id"],
            "email": user["email"],
            "token": token,
            "created_at": datetime.now(timezone.utc),
            "expires_at": expires_at,
            "used": False,
        }},
        upsert=True,
    )
    origin = request.headers.get("origin") or str(request.base_url).rstrip("/")
    verify_url = f"{origin}/verify-email?token={token}"
    sent = await _send_verification_email(user["email"], verify_url)
    resp: Dict[str, Any] = {"ok": True, "message": "Email de verificación enviado"}
    if not sent:
        resp["dev_token"] = token
        resp["dev_verify_url"] = verify_url
    return resp


@router.post("/auth/verify-email")
async def verify_email(payload: VerifyEmailRequest):
    """Confirm email ownership using the token from the verification email."""
    rec = await db.email_verification_tokens.find_one({"token": payload.token, "used": False})
    if not rec:
        raise HTTPException(status_code=400, detail="Token inválido o ya utilizado")
    expires_at = rec["expires_at"]
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at.replace("Z", "+00:00"))
    if expires_at.tzinfo is None:
        expires_at = expires_at.replace(tzinfo=timezone.utc)
    if datetime.now(timezone.utc) > expires_at:
        raise HTTPException(status_code=400, detail="El token ha expirado. Solicita uno nuevo.")

    await db.users.update_one(
        {"id": rec["user_id"]},
        {"$set": {"email_verified": True, "email_verified_at": datetime.now(timezone.utc).isoformat()}},
    )
    await db.email_verification_tokens.update_one(
        {"token": payload.token},
        {"$set": {"used": True, "used_at": datetime.now(timezone.utc)}},
    )
    return {"ok": True, "message": "Email verificado correctamente"}


# ---------------------------------------------------------------------------
# 8.  SUBSCRIPTIONS — REAL change-plan (Stripe proration)
# ---------------------------------------------------------------------------

class ChangePlanRequest(BaseModel):
    new_plan_id: str
    proration_behavior: str = "create_prorations"  # or "none"


@router.post("/subscriptions/change-plan")
async def change_plan_real(payload: ChangePlanRequest, user: dict = Depends(_require_user_proxy)):
    """
    Upgrade or downgrade an active Stripe subscription via proration.
    Requires the user to have a Stripe customer + subscription.
    If neither exists, redirects to checkout.
    """
    if payload.new_plan_id not in SUBSCRIPTION_PLANS:
        raise HTTPException(status_code=400, detail="Plan inválido")

    user_doc = await db.users.find_one({"id": user["id"]}, {"_id": 0})
    if not user_doc:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    # Runtime Stripe key
    runtime_key = (await get_setting("stripe_secret_key")) or STRIPE_API_KEY
    stripe.api_key = runtime_key

    customer_id = user_doc.get("stripe_customer_id")
    subscription_id = user_doc.get("stripe_subscription_id")

    if not customer_id or not subscription_id:
        return {
            "ok": False,
            "redirect_to_checkout": True,
            "message": "No hay suscripción activa en Stripe. Realiza una nueva compra.",
            "new_plan_id": payload.new_plan_id,
        }

    try:
        sub = stripe.Subscription.retrieve(subscription_id)
        if sub.status not in ("active", "trialing"):
            return {
                "ok": False,
                "redirect_to_checkout": True,
                "message": f"La suscripción actual no está activa (estado: {sub.status}). Realiza una nueva compra.",
            }

        # Get current subscription item id
        item_id = sub["items"]["data"][0]["id"]
        new_plan = SUBSCRIPTION_PLANS[payload.new_plan_id]

        # We need a Stripe Price ID for the new plan.
        # Look it up from our app_settings or create a one-off price.
        price_key = f"stripe_price_{payload.new_plan_id}"
        stripe_price_id = await get_setting(price_key)

        if not stripe_price_id:
            # Create an ad-hoc price for this plan
            price_obj = stripe.Price.create(
                currency=new_plan["currency"].lower(),
                unit_amount=int(new_plan["price"] * 100),
                recurring={"interval": new_plan.get("stripe_interval", "month")},
                product_data={"name": f"Trading Calculator PRO — {new_plan['name']}"},
            )
            stripe_price_id = price_obj.id

        updated_sub = stripe.Subscription.modify(
            subscription_id,
            items=[{"id": item_id, "price": stripe_price_id}],
            proration_behavior=payload.proration_behavior,
        )

        # Update our DB
        new_end = datetime.now(timezone.utc) + timedelta(days=new_plan["days"])
        await db.users.update_one(
            {"id": user["id"]},
            {"$set": {
                "subscription_plan": payload.new_plan_id,
                "subscription_end": new_end.isoformat(),
                "is_premium": True,
            }},
        )
        return {
            "ok": True,
            "message": f"Plan cambiado a {new_plan['name']} correctamente.",
            "new_plan_id": payload.new_plan_id,
            "subscription_status": updated_sub.status,
            "proration_behavior": payload.proration_behavior,
        }
    except stripe.error.StripeError as e:
        raise HTTPException(status_code=400, detail=f"Error de Stripe: {e.user_message or str(e)}")
    except Exception as e:
        logging.error(f"change_plan_real error: {e}")
        raise HTTPException(status_code=500, detail="Error cambiando plan")


# ---------------------------------------------------------------------------
# 9.  STRIPE WEBHOOK — subscription.deleted + invoice.payment_failed
#     (This is an ADDITIONAL handler; the main checkout.session.completed is
#      still handled by the existing /webhook/stripe in server.py)
# ---------------------------------------------------------------------------

@router.post("/webhook/stripe/subscription")
async def stripe_subscription_webhook(request: Request) -> Dict[str, str]:
    """
    Handle Stripe subscription lifecycle events:
    - customer.subscription.deleted  → deactivate user premium
    - invoice.payment_failed         → mark subscription past_due
    - customer.subscription.updated  → sync status changes
    """
    payload_bytes = await request.body()
    sig = request.headers.get("Stripe-Signature", "")

    runtime_key = (await get_setting("stripe_secret_key")) or STRIPE_API_KEY
    webhook_secret = (await get_setting("stripe_webhook_secret")) or os.environ.get("STRIPE_WEBHOOK_SECRET", "")
    stripe.api_key = runtime_key

    if not webhook_secret:
        raise HTTPException(status_code=400, detail="Stripe webhook secret not configured")

    try:
        event = stripe.Webhook.construct_event(payload_bytes, sig, webhook_secret)
    except stripe.error.SignatureVerificationError:
        raise HTTPException(status_code=400, detail="Invalid Stripe signature")
    except Exception as e:
        logging.error(f"Stripe webhook parse error: {e}")
        raise HTTPException(status_code=400, detail="Bad webhook payload")

    event_type = event["type"]
    data_obj = event["data"]["object"]

    if event_type == "customer.subscription.deleted":
        customer_id = data_obj.get("customer")
        if customer_id:
            user = await db.users.find_one({"stripe_customer_id": customer_id}, {"_id": 0})
            if user:
                await db.users.update_one(
                    {"stripe_customer_id": customer_id},
                    {"$set": {
                        "is_premium": False,
                        "subscription_plan": None,
                        "subscription_end": None,
                        "subscription_status": "canceled",
                        "stripe_subscription_id": None,
                        "subscription_canceled_at": datetime.now(timezone.utc).isoformat(),
                    }},
                )
                logging.info(f"Subscription deleted for customer {customer_id} ({user.get('email')})")

    elif event_type == "invoice.payment_failed":
        customer_id = data_obj.get("customer")
        attempt = data_obj.get("attempt_count", 1)
        if customer_id:
            update: Dict[str, Any] = {"subscription_status": "past_due"}
            # After 3 failed attempts, revoke premium access
            if attempt and int(attempt) >= 3:
                update["is_premium"] = False
                update["subscription_plan"] = None
                update["subscription_status"] = "unpaid"
                logging.warning(f"Payment failed {attempt}x for {customer_id} — premium revoked")
            await db.users.update_one(
                {"stripe_customer_id": customer_id},
                {"$set": update},
            )

    elif event_type == "customer.subscription.updated":
        customer_id = data_obj.get("customer")
        status = data_obj.get("status")
        if customer_id and status:
            is_active = status in ("active", "trialing")
            await db.users.update_one(
                {"stripe_customer_id": customer_id},
                {"$set": {"subscription_status": status, "is_premium": is_active}},
            )

    return {"status": "received", "event": event_type}


# ---------------------------------------------------------------------------
# 10.  PERFORMANCE EXPORT (CSV / Excel)
# ---------------------------------------------------------------------------

_TRADE_EXPORT_FIELDS = [
    "id", "symbol", "side", "setup", "entry_price", "exit_price",
    "sl", "tp", "quantity", "pnl", "pnl_pct", "roe", "fees",
    "entry_date", "exit_date", "status", "notes", "tags", "emotion",
]


def _trades_to_csv_bytes(trades: List[Dict[str, Any]]) -> bytes:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=_TRADE_EXPORT_FIELDS, extrasaction="ignore")
    writer.writeheader()
    for t in trades:
        row = {k: t.get(k, "") for k in _TRADE_EXPORT_FIELDS}
        if isinstance(row.get("tags"), list):
            row["tags"] = ",".join(row["tags"])
        writer.writerow(row)
    return buf.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility


def _trades_to_excel_bytes(trades: List[Dict[str, Any]]) -> bytes:
    """Generate an Excel file. Uses openpyxl if available, falls back to CSV."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trades"
        # Header row with styling
        header_fill = PatternFill("solid", fgColor="004C97")
        for col_idx, field in enumerate(_TRADE_EXPORT_FIELDS, start=1):
            cell = ws.cell(row=1, column=col_idx, value=field.upper())
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        # Data rows
        for row_idx, trade in enumerate(trades, start=2):
            for col_idx, field in enumerate(_TRADE_EXPORT_FIELDS, start=1):
                val = trade.get(field, "")
                if isinstance(val, list):
                    val = ",".join(str(v) for v in val)
                ws.cell(row=row_idx, column=col_idx, value=val)
        buf_bytes = io.BytesIO()
        wb.save(buf_bytes)
        return buf_bytes.getvalue()
    except ImportError:
        # openpyxl not installed, return CSV
        return _trades_to_csv_bytes(trades)


@router.get("/performance/export")
async def export_trades(
    user: dict = Depends(_require_user_proxy),
    format: str = "csv",   # "csv" or "excel"
    status: Optional[str] = None,
    symbol: Optional[str] = None,
    since: Optional[str] = None,
    until: Optional[str] = None,
):
    """
    Export all trades to CSV or Excel.
    Query params: format=csv|excel, status=open|closed, symbol=AAPL, since=2024-01-01, until=2024-12-31
    """
    query: Dict[str, Any] = {"user_id": user["id"]}
    if status:
        query["status"] = status
    if symbol:
        query["symbol"] = symbol.upper()
    if since or until:
        date_filter: Dict[str, Any] = {}
        if since:
            date_filter["$gte"] = since
        if until:
            date_filter["$lte"] = until
        query["entry_date"] = date_filter

    cursor = db.trades.find(query, {"_id": 0}).sort("entry_date", -1)
    trades = await cursor.to_list(length=10000)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    if format.lower() == "excel":
        content = _trades_to_excel_bytes(trades)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        filename = f"trades_{ts}.xlsx"
    else:
        content = _trades_to_csv_bytes(trades)
        media_type = "text/csv; charset=utf-8"
        filename = f"trades_{ts}.csv"

    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ---------------------------------------------------------------------------
# 11.  SAVE CALCULATION TO JOURNAL
# ---------------------------------------------------------------------------

class SaveToJournalRequest(BaseModel):
    symbol: str = ""
    direction: str = "long"       # long | short
    entry_price: Optional[float] = None
    quantity: Optional[float] = None
    stop_loss: Optional[float] = None
    take_profit: Optional[float] = None
    notes: Optional[str] = None
    tags: Optional[List[str]] = None


@router.post("/calculations/{calc_id}/save-to-journal")
async def save_calculation_to_journal(calc_id: str, payload: SaveToJournalRequest, user: dict = Depends(_require_user_proxy)):
    """
    Pre-fill a new journal trade entry from a saved calculation.
    The calculation inputs/results are merged with any overrides in the payload.
    """
    calc = await db.calculations.find_one({"id": calc_id, "user_id": user["id"]}, {"_id": 0})
    if not calc:
        raise HTTPException(status_code=404, detail="Cálculo no encontrado")

    inputs = calc.get("inputs", {})
    results = calc.get("results", {})

    # Merge: explicit payload fields override inferred from calculation
    symbol = payload.symbol or inputs.get("symbol", "UNKNOWN")
    direction = payload.direction or inputs.get("direction", "long")
    entry_price = payload.entry_price or inputs.get("entry_price") or inputs.get("entryPrice") or 0.0
    quantity = payload.quantity or inputs.get("quantity") or inputs.get("positionSize") or 1.0
    stop_loss = payload.stop_loss or inputs.get("stop_loss") or inputs.get("stopLoss")
    take_profit = payload.take_profit or inputs.get("take_profit") or inputs.get("takeProfit")
    notes_base = payload.notes or ""
    auto_note = (
        f"[Auto] Creado desde cálculo: {calc.get('calculator_type', 'calculadora')} — "
        f"Resultado: {results}"
    )
    notes = f"{notes_base}\n{auto_note}".strip() if notes_base else auto_note

    trade_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "symbol": symbol.upper(),
        "side": direction,
        # Journal v2 fields (performance module)
        "entry_price": float(entry_price),
        "exit_price": None,
        "sl": float(stop_loss) if stop_loss else None,
        "tp": float(take_profit) if take_profit else None,
        "quantity": float(quantity),
        "status": "open",
        "notes": notes,
        "tags": payload.tags or [f"calc:{calc.get('calculator_type', 'calc')}"],
        "entry_date": datetime.now(timezone.utc).isoformat(),
        "exit_date": None,
        "fees": 0.0,
        "emotion": None,
        "account_balance": 0.0,
        "source_calculation_id": calc_id,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.trades.insert_one(trade_doc)
    trade_doc.pop("_id", None)
    return {
        "ok": True,
        "trade_id": trade_doc["id"],
        "message": f"Trade pre-rellenado desde cálculo y añadido al diario.",
        "trade": trade_doc,
    }


# ---------------------------------------------------------------------------
# TTL indexes for new collections added in this module
# ---------------------------------------------------------------------------

async def ensure_missing_api_indexes(database) -> None:
    """Create TTL and query indexes for new collections."""
    try:
        # password_reset_tokens: auto-expire 2h after expiry
        await database.password_reset_tokens.create_index(
            "expires_at", expireAfterSeconds=7200, name="prt_ttl",
        )
        await database.password_reset_tokens.create_index("token", unique=True, sparse=True)
        await database.password_reset_tokens.create_index("user_id")

        # email_verification_tokens
        await database.email_verification_tokens.create_index(
            "expires_at", expireAfterSeconds=86400, name="evt_ttl",
        )
        await database.email_verification_tokens.create_index("token", unique=True, sparse=True)
        await database.email_verification_tokens.create_index("user_id")

        # Unique email index
        await ensure_email_unique_index(database)

        logging.info("✅ missing_apis indexes ensured")
    except Exception as e:
        logging.error(f"missing_apis index error: {e}")


# ---------------------------------------------------------------------------
# REGISTRATION HELPER — called from server.py startup
# ---------------------------------------------------------------------------

def register(
    app_router: APIRouter,
    database,
    helpers: Dict[str, Any],
) -> None:
    """
    Bind this module's router to the main app router and inject shared
    dependencies (db, auth helpers, config) so we avoid circular imports.

    Call from server.py after creating `api_router`:

        from missing_apis import router as missing_router, register as register_missing
        register_missing(api_router, db, {
            "require_user": require_user,
            "check_premium": check_premium,
            "SUBSCRIPTION_PLANS": SUBSCRIPTION_PLANS,
            "hash_password": hash_password,
            "SENDGRID_API_KEY": SENDGRID_API_KEY,
            "SENDER_EMAIL": SENDER_EMAIL,
            "STRIPE_API_KEY": STRIPE_API_KEY,
            "get_setting": get_setting,
        })
    """
    global db, require_user, check_premium, SUBSCRIPTION_PLANS
    global hash_password, SENDGRID_API_KEY, SENDER_EMAIL, STRIPE_API_KEY, get_setting

    db = database
    require_user = helpers["require_user"]
    check_premium = helpers["check_premium"]
    SUBSCRIPTION_PLANS = helpers["SUBSCRIPTION_PLANS"]
    hash_password = helpers["hash_password"]
    SENDGRID_API_KEY = helpers["SENDGRID_API_KEY"]
    SENDER_EMAIL = helpers["SENDER_EMAIL"]
    STRIPE_API_KEY = helpers["STRIPE_API_KEY"]
    get_setting = helpers["get_setting"]

    app_router.include_router(router)
